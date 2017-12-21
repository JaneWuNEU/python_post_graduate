# -*- coding: utf-8 -*-
import numpy
import os
from openpyxl import load_workbook 
from openpyxl.workbook import Workbook

import matplotlib.pyplot as plt
class WRFile:
    '''
the data produced by this method fit normal distribution,
and the mean ,var ,and quantity are given by the user
    '''
   

    """
the argument must be an array,
and the write it into excel
    """
    def writeDataIntoExcel(self,data,filePath,describe ="no description",cols = 0,sheet_name = "1"):
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        try:
            colNum = len(data[0]) 
            print(colNum)
            for i in range(len(data)):
               temp = data[i]
               for j in range(colNum):
                   ws.cell(row = i+1,column = j+1,value = temp[j])
        except TypeError:
            colNum = 1
            print("here")
            for i in range(len(data)):
               temp = data[i]
               ws.cell(row = i+1,column = colNum,value = temp)
        wb.save(filePath)
        
    def writeDataIntoExcelLList(self,data,filePath,sheet_name="1"):
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        for i in range(len(data)):
            row = i+1
            row_data = data[i]
            try:
                col_max = len(row_data)
                for j in range(col_max):
                    ws.cell(row = row,column = j+1,value = data[i][j])
            except TypeError:
                col_max = 1
                ws.cell(row =row,column = col_max,value = row_data)
        wb.save(filePath)
    
    '''
@filePath  the source of the data
@result read the data, put it into the array,
        and return to the user
    '''
    def readDataFromExcel(self,filePath,cols=1,sheet_name = "1",min_cols = 1,max_cols = 1):
        result = []
        
        if os.path.exists(filePath):
            #second,read the data into the workbook,then get the sheet
            wb = load_workbook(filename = filePath)
            ws = wb.get_sheet_by_name(sheet_name)
            
            for col in ws.iter_cols(min_col=min_cols, max_col=max_cols):                 
                for cell in col:
                    result.append(float(cell.value))
            result = numpy.array(result) 
        
        return result
        
    def readDataFromExcel2(self,filePath,rows = 1,sheet_name = "1",Min_rows = None,Max_rows = None):

        result = []
        #fisrt,to check whether the file exists
        if os.path.exists(filePath):
            #second,read the data into the workbook,then get the sheet
            wb = load_workbook(filename = filePath)
            ws = wb.get_sheet_by_name(sheet_name)
            if Min_rows==None:
               Min_rows = ws.min_row
            if Max_rows == None:
               Max_rows = ws.max_row
            for row in ws.iter_rows(min_row=Min_rows, max_row=Max_rows):                
                row_cell_value=[]
                for cell in row:
                    row_cell_value.append(cell.value)   
                result.append(row_cell_value)
            
           
        return result
      
     

   
        
        
        
        
        
        
        
        
        
        
        