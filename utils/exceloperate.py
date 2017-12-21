
#create a workbook
from openpyxl import Workbook
wb = Workbook()
#get the worksheet
ws = wb.active

'''operate the sheet'''
#create a worksheet
sheet_location = 1
ws1 = wb.create_sheet("sheet_name",sheet_location)
ws1.title("change the sheet name")

#once you change the sheet's name,you can use the new name as a key to get it
ws3 = wb["change the sheet name"]

#loop through worksheets,and the wb is a tuple
for sheet in wb:
    print(sheet.title)
#we can create a copies of worksheets within a single workbook
source = wb.active
target = wb.copy_worksheet(source)

'''
accessing the cell
all the cells are created in the memory
only when you save them can they turn into a real file stored in the desk
'''
#get the value of the cell A4
c = ws['A4']
#change the value of the cell A4
ws['A4'] = "WJ"
#give and row and col to get the value
d = ws.cell(row = 4,column = 2,value = 10)
'''
important operation
what's the data type of the result tuple or array
'''
cell_range = ws['A1':'C2']
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

for row in ws.iter_rows(min_row = 1,max_col = 3,max_row = 2):
    for cell in row:
        print(cell)

for col in ws.iter_cols(min_col = 1,max_col = 3,max_row = 2):
    for cell in col:
        print(cell)

ws = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
wb.save(fileName)












