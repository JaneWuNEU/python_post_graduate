# -*- coding: utf-8 -*-
import numpy
import os
from openpyxl import load_workbook 
from openpyxl.workbook import Workbook
import matplotlib.pyplot as plt
import sys
sys.path.append("D:\\anaconda\\project")
from plotdata import PlotData

class WRFile:

    '''
the data produced by this method fit normal distribution,
and the mean ,var ,and quantity are given by the user
    '''
   

    """
the argument must be an array,
and the write it into excel
    """
    def writeDataIntoExcel(self,data,filePath,describe ="no description",cols = 0,sheet_name = "1"):
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        try:
            colNum = len(data[0]) 
            print(colNum)
            for i in range(len(data)):
               temp = data[i]
               for j in range(colNum):
                   ws.cell(row = i+1,column = j+1,value = temp[j])
        except TypeError:
            colNum = 1
            
            for i in range(len(data)):
               temp = data[i]
               ws.cell(row = i+1,column = colNum,value = temp)
        wb.save(filePath)
        
    
    '''
@filePath  the source of the data
@result read the data, put it into the array,
        and return to the user
    '''
    def readDataFromExcel(self,filePath,cols=1,sheet_name = "1",min_cols = 1,max_cols = 1):
        result = []
        
        if os.path.exists(filePath):
            #second,read the data into the workbook,then get the sheet
            wb = load_workbook(filename = filePath)
            ws = wb.get_sheet_by_name(sheet_name)
            
            for col in ws.iter_cols(min_col=min_cols, max_col=max_cols):                 
                for cell in col:
                    result.append(float(cell.value))
            result = numpy.array(result) 
        
        return result
        
    def readDataFromExcel2(self,filePath,rows = 1,sheet_name = "1",min_row =1,max_row = 1):
        result = []
        #fisrt,to check whether the file exists
        if os.path.exists(filePath):
            #second,read the data into the workbook,then get the sheet
            wb = load_workbook(filename = filePath)
            ws = wb.get_sheet_by_name(sheet_name)
            
            for row in ws.iter_rows(min_row=min_row, max_row=max_row):                
                row_cell_value=[]
                for cell in row:
                    row_cell_value.append(cell.value)   
                result.append(row_cell_value)
            numpy.array(result)
           
        return result
'''    
wrFile = WRFile()
data = wrFile.readDataFromExcel(filePath = "F:\\baidu\\online/taobao.xlsx",sheet_name = "Sheet1",min_cols = 2,max_cols = 2)
print(numpy.percentile(data,50))
print(numpy.percentile(data,90)) 
print(numpy.max(data))
print(numpy.min(data))
'''        
        