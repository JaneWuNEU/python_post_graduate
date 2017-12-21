# -*- coding: utf-8 -*-
"""
Created on Sun Jan 15 15:13:30 2017

@author: User
"""
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
def dataList():
    dataList =[]

    begin = "wc_day"
    
    subFile = [3,8,5,5,5,4]
    mainFile = np.arange(50,56).tolist()
    print(mainFile)
    time3 = []
    for i in range(0,6):
        k = 0
        while k<subFile[i]:
            time3.append(str(mainFile[i])+"_"+str(k+2))      
            k+=1
    fileName = begin
    for j in range(0,len(time3)):
        fileName = fileName+time3[j]
        dataList.append(fileName)
        fileName = begin
    return dataList
def getTimeModel():
        data = np.linspace(0,235959,235959).tolist()
        result = []
        

        for i in range(0,len(data)):
            point = str(int(data[i]))
            point_len = len(point)
            left_pos = 6-point_len#左侧补位
            j = 0
            left_con =""
            while j<left_pos:
                left_con+="0"
                j+=1
            point = left_con+point
            print(point)
            result.append(point)
        return  result
         
def getWorkloadofDay():        
    filePath = "F:/one/"

    #获取文件列表
    fileList = ["wc_day50_1"]
    #获取时间点
    #timePoints = getTimeModel()
    
    for i in range(0,len(fileList)):
        workload = np.zeros(235959+1)
        fileObject = open("D:\\virtualbox\\share\\data/"+fileList[i]+".out",'r',encoding = "utf-8")
        while True:
            try:   
               row = fileObject.readline()
               if not row:
                   break
               else:
                   local = row.find(":")+1
                   temp = row[local:local+8]
                   temp = temp.replace(":","")
                   print(temp)
                   temp = int(temp)
                   workload[temp]+=1
            except UnicodeDecodeError:
                   pass
            except ValueError:
                    pass
                   #print(temp+" 值异常")
        fileObject.close()#关闭当前文件夹
    
        #确定输出文件的名称
        #fileList[i] = fileList[i][0:fileList[i].rfind(".")]
        #print(fileList[i])
        writeDataIntoExcel(workload.tolist(),"F:/one/output/"+fileList[i]+".xlsx",nozero = False)
def writeDataIntoExcel(workload,filePath,nozero = False):
        sheet_name = "1"
        cols = 1
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        k=1
        for i in range(len(workload)):
            '''
            if workload[i]==0:
                if nozero:
                    continue
            '''
            ws.cell(row = i+1,column = cols,value =workload[i])
            
        wb.save(filePath)
def readDataFromExcel(filePath):
        result = []
        sheet_name = 1
        cols = 1
        #fisrt,to check whether the file exists
        if os.path.exists(filePath):
            #second,read the data into the workbook,then get the sheet
            wb = load_workbook(filename = filePath)
            
            ws = wb.get_sheet_by_name("1")
            for row in ws.iter_rows():
                for cell in row:
                    result.append(cell.value)      
            result = np.array(result,dtype = np.float)
            return result      
def combineWorkload():
    fileList = os.listdir("F:/one/output")
    days = np.zeros(92)
    for i in range(0,len(fileList)):
        fileName = fileList[i]
        print(fileName[6:8])
        day = int(fileName[6:8])
        days[day-1]+=1#统计每天有多少个子文件
        
        #fileObject = open("F:/one/output"+fileName,'r',encoding = "utf-8")
        #获取day相同的文件目录
    
    for j in range(0,92):
        subFiles = days[j]#第（j+1）天有days[j]个子文件
        if subFiles==0:
            continue
        workload = readDataFromExcel("F:/one/output/wc_day"+str(j+1)+"_1.xlsx")
        if subFiles>=2:
            for k in range(2,int(subFiles)+1):#子文件的下标从1开始
                '''为了节省内存空间，对子文件进行两两处理'''
                temp = readDataFromExcel("F:/one/output/wc_day"+str(j+1)+"_"+str(k)+".xlsx")
                #print(len(temp))
                workload+=temp
                
        """把最终的文件写入text文件"""
        wb = Workbook()
        ws = wb.create_sheet("1")
        k=1
        for i in range(len(workload)):
            if i%100>59:
                continue
            elif i%10000>5959:
                continue
            else: 
                ws.cell(row = k,column = 1,value = workload[i])
                k+=1
        wb.save("F:/one/total/wc_day"+str(j+1)+".xlsx")        

def getFinalWorkload():
    
    
    part0 = "F:/one/total/wc_day"
    part1 = ".xlsx"
    wb = Workbook()
    ws = wb.create_sheet("1")
    k=1
    for i in range(50,57):
       workload = readDataFromExcel(part0+str(i)+part1)
       for j in range(len(workload)):
          ws.cell(row = k,column = 1,value = workload[j])
          k+=1 
    wb.save("F:/one/final/wc_day50_56.xlsx")

def processFinalWorkload():
    workload = readDataFromExcel("F:/one/total/wc_day51.xlsx")
    temp = readDataFromExcel("F:/one/total/wc_day52.xlsx")
    workload+=temp
    wb = Workbook()
    ws = wb.create_sheet("1")
    k=1
    for i in range(len(workload)):
        if i%100>59:
            continue
        elif i%10000>5959:
            continue
        else: 
            ws.cell(row = k,column = 1,value = workload[i])
            k+=1
    wb.save("F:/one/final/workload51_52.xlsx")
def combineDataInMinute():    
    data = readDataFromExcel("F:\\FIFA\\worldcup\\workload58_62.xlsx")
    counts = len(data)

getWorkloadofDay()    
#getWorkloadofDay()
#combineWorkload()

#data = readDataFromExcel("F:/one/final/wc_day51_56.xlsx")
